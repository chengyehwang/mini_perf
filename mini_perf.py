#!/usr/bin/env python3
import numpy as np
import sys
import pandas as pd
import re
import glob
import simplejson as json
import argparse

l1_penalty = 0
l2_penalty = 10 # 5ns
l3_penalty = 40 # 20ns
dram_penalty = 200 # 100ns

parser = argparse.ArgumentParser()
parser.add_argument("--last", action="store_true", help="handle last data")
args = parser.parse_args()

name = []
filename = glob.glob('*mini_perf.head')[0]
with open(filename, 'r') as f:
    for line in f.readlines():
        m = re.search('count:\s*(\S+)',line)
        if m:
            name.append(m.group(1))
#print(name)
def read_file():
    if args.last:
        file_data = filename.replace('.head','.last_data')
    else:
        file_data = filename.replace('.head','.data')
    array = []
    with open(file_data, 'rb') as f:
        data = np.fromfile(f, dtype='int64')
        nx = len(name)
        ny = int(len(data)/nx)
        #print('matrix', nx, ny, len(data))
        array = np.reshape(data, [ny, nx])
    return array

counter = read_file()
#print(counter)
data = pd.DataFrame(counter,columns=name)

data['delta'] = data['time']
data['time'] = data['time'] / 1000000000.0
data = data.set_index('time')

data = data.diff().dropna()

# group update running / enabled
for counter in data.columns:
    #print(counter)
    #print(data.columns)
    m = re.search('^(.*)(_g\d)(_cpu\d)$', counter)
    if m:
        counter_new = m.group(1) + m.group(3)
        post_fix = m.group(2) + m.group(3)
        if counter_new in data.columns:
            continue
        data[counter_new] = (data[counter] * data['delta'] / data['time_running'+post_fix]).round()

data = data.fillna(0)

if False: # Q 4ms + pmu timeinterleave -> moving average
    data = data.rolling(window=6).mean()

cpu_list = ['0','1','2','3','4','5','6','7']
cpu_group = {'L': ['0','1','2','3'], 'B': ['4','5','6','7']}

count_list = set()
for i in data.columns:
    m = re.search('^(.*)_cpu\d$',i)
    if m:
        count_list.add(m[1])

for count in count_list:
    for group in cpu_group:
        data[count + '_cpu' + group] = 0
        for cpu in cpu_group[group]:
            data[count + '_cpu' + group] += data[count + '_cpu' + cpu]

for group in cpu_group:
    cpu_list.append(group)

if False:
    print(data.columns)

    for index, row in data.iterrows():
        for column in row.keys():
            print(column, row[column])


for cpu in cpu_list:
    cpu = '_cpu' + cpu
    for group in range(8):
        group = '_g%d' % group
        for cache in ['l1i', 'l1d', 'l2d', 'l3d']:
            count_refill = 'raw-' + cache + '-cache-refill' + group + cpu
            count_total = 'raw-' + cache + '-cache' + group + cpu
            count_miss = cache + '-cache-miss' + cpu
            if count_refill in data.columns and count_total in data.columns:
                data[count_miss] = data[count_refill] / data[count_total]

        count_mips = 'mips' + cpu

        count_time = 'time_running' + group + cpu
        count_inst = 'raw-inst-retired' + group + cpu
        count_i = 'raw-l1i-cache' + group + cpu
        count_d = 'raw-l1d-cache' + group + cpu
        count_i_req = 'l1i-cache-req' + cpu
        count_d_req = 'l1d-cache-req' + cpu

        if count_inst in data.columns:
            data[count_mips] = (data[count_inst] / 1000000) / (data[count_time] / 1000000000)

        if count_inst in data.columns and count_i in data.columns:
            data[count_i_req] = data[count_i] / data[count_inst]
        if count_inst in data.columns and count_d in data.columns:
            data[count_d_req] = data[count_d] / data[count_inst]

        count_cycle = 'raw-cpu-cycles' + group + cpu
        cpi = 'cpi' + cpu
        if count_inst in data.columns and count_cycle in data.columns:
            data[cpi] = data[count_cycle] / data[count_inst]

for cpu in cpu_list:
    cpu = '_cpu' + cpu
    l1i_miss = 'l1i-cache-miss' + cpu
    l1d_miss = 'l1d-cache-miss' + cpu
    l2_miss = 'l2d-cache-miss' + cpu
    l3_miss = 'l3d-cache-miss' + cpu
    l1_hit = 'l1-cache-hit' + cpu
    l2_hit = 'l2-cache-hit' + cpu
    l3_hit = 'l3-cache-hit' + cpu
    dram_hit = 'dram-hit' + cpu
    l1i_req = 'l1i-cache-req' + cpu
    l1d_req = 'l1d-cache-req' + cpu
    l2_req = 'l2-cache-req' + cpu
    l3_req = 'l3-cache-req' + cpu

    if l1i_req in data.columns and l1d_req in data.columns and l1i_miss in data.columns and l1d_miss in data.columns:
        data[l1_hit] = 1 - ( data[l1i_req] * data[l1i_miss] + data[l1d_req] * data[l1d_miss] )
        data[l2_req] = 1 - data[l1_hit]

    if l2_req in data.columns and l2_miss in data.columns:
        data[l2_hit] = data[l2_req] * (1 - data[l2_miss])
        data[l3_req] = data[l2_req] * data[l2_miss]

    if l3_req in data.columns and l3_miss in data.columns:
        data[l3_hit] = data[l3_req] * (1 - data[l3_miss])
        data[dram_hit] = data[l3_req] * data[l3_miss]

    if l1_hit in data.columns and l2_hit in data.columns and l3_hit in data.columns and dram_hit in data.columns:
        impact = 'cache_impact' + cpu
        data[impact] = data[l1_hit] * l1_penalty + data[l2_hit] * l2_penalty + data[l3_hit] * l3_penalty + data[dram_hit] * dram_penalty;

columns = list(data.columns)
def compare(item):
    return item[-1]

columns = sorted(columns, key=compare)

data = data[columns]

if args.last:
    file_json = filename.replace('.head','.json')
    with open(file_json, 'w') as fp:
        result = data.iloc[-1].to_json()
        parsed = json.loads(result)
        fp.write(json.dumps(parsed, indent=4))
    sys.exit(0)

#print(data)
file_excel = filename.replace('.head','.xlsx')
data.to_excel(file_excel, sheet_name = 'pmu')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import PatternFill

workbook = load_workbook(file_excel)
workbook['pmu'].freeze_panes = "B2"
workspace = workbook.create_sheet(title = 'impact')

if True: # time
    for row in range(1, len(data.index) + 2):
        workspace['A' + str(row)] = '=pmu!' + 'A' + str(row)

column = 2

for cpu in cpu_list:
    cpu = '_cpu' + cpu
    cpi = 'cpi' + cpu
    mips = 'mips' + cpu
    l1_hit = 'l1-cache-hit' + cpu
    l2_hit = 'l2-cache-hit' + cpu
    l3_hit = 'l3-cache-hit' + cpu
    dram_hit = 'dram-hit' + cpu
    CPI = -1
    MIPS = -1
    L1 = -1
    L2 = -1
    L3 = -1
    L3 = -1
    for i in range(len(data.columns)): # index as first column = 1
        if cpi == data.columns[i]:
            CPI = i + 2

        if mips == data.columns[i]:
            MIPS = i + 2

        if l1_hit == data.columns[i]:
            L1 = i + 2

        if l2_hit == data.columns[i]:
            L2 = i + 2

        if l3_hit == data.columns[i]:
            L3 = i + 2

        if dram_hit == data.columns[i]:
            L4 = i + 2
    if CPI < 0 or MIPS < 0 or L1 < 0 or L2 < 0 or L3 < 0 or L4 < 0:
        continue
    # penalty columns
    w0 = get_column_letter(column+1) + '1'
    w1 = get_column_letter(column+1) + '2'
    w2 = get_column_letter(column+1) + '3'
    w3 = get_column_letter(column+1) + '4'
    w4 = get_column_letter(column+1) + '5'
    w5 = get_column_letter(column+1) + '6'
    w21 = get_column_letter(column+1) + '22'
    w37 = get_column_letter(column+1) + '38'
    workspace[w0] = 'penalty' + cpu
    workspace[w1] = l1_penalty 
    workspace[w2] = l2_penalty # 5ns
    workspace[w3] = l3_penalty # 20ns
    workspace[w4] = dram_penalty # 100ns

    workspace[w1].fill = PatternFill(start_color="FFEE08", end_color="FFEE08", fill_type = "solid")
    workspace[w2].fill = PatternFill(start_color="FFEE08", end_color="FFEE08", fill_type = "solid")
    workspace[w3].fill = PatternFill(start_color="FFEE08", end_color="FFEE08", fill_type = "solid")
    workspace[w4].fill = PatternFill(start_color="FFEE08", end_color="FFEE08", fill_type = "solid")

    row_num = len(data.index) + 1

    # hit
    for row in range(1, row_num + 1):
        impact = get_column_letter(column) + str(row)
        cpi = get_column_letter(column+2) + str(row)
        mips = get_column_letter(column+3) + str(row)
        l1 = get_column_letter(column+4) + str(row)
        l2 = get_column_letter(column+5) + str(row)
        l3 = get_column_letter(column+6) + str(row)
        l4 = get_column_letter(column+7) + str(row)
        if row ==1:
            workspace[impact] = "cache_impact" + cpu
        else:
            workspace[impact] = "= %s * %s + %s * %s + %s * %s + %s * %s"%(l1,w1,l2,w2,l3,w3,l4,w4)
        workspace[cpi] = "=pmu!" + get_column_letter(CPI) + str(row)
        workspace[mips] = "=pmu!" + get_column_letter(MIPS) + str(row)
        workspace[l1] = "=pmu!" + get_column_letter(L1) + str(row)
        workspace[l2] = "=pmu!" + get_column_letter(L2) + str(row)
        workspace[l3] = "=pmu!" + get_column_letter(L3) + str(row)
        workspace[l4] = "=pmu!" + get_column_letter(L4) + str(row)

    # chart
    chart = ScatterChart()
    chart.title = 'Cache Impact' + cpu
    chart.x_axis.title = 'Time'
    xvalues = Reference(workspace, min_col = 1, min_row = 2, max_row = row_num + 1)
    values = Reference(workspace, min_col = column, min_row = 2, max_row = row_num + 1)
    series = Series(values, xvalues)
    chart.series.append(series)
    chart.legend = None
    workspace.add_chart(chart, w5)

    chart = ScatterChart()
    chart.title = 'CPI' + cpu
    chart.x_axis.title = 'Time'
    xvalues = Reference(workspace, min_col = 1, min_row = 2, max_row = row_num + 1)
    values = Reference(workspace, min_col = column+2, min_row = 2, max_row = row_num + 1)
    series = Series(values, xvalues)
    chart.series.append(series)
    chart.legend = None
    workspace.add_chart(chart, w21)

    chart = ScatterChart()
    chart.title = 'MIPS' + cpu
    chart.x_axis.title = 'Time'
    xvalues = Reference(workspace, min_col = 1, min_row = 2, max_row = row_num + 1)
    values = Reference(workspace, min_col = column+3, min_row = 2, max_row = row_num + 1)
    series = Series(values, xvalues)
    chart.series.append(series)
    chart.legend = None
    workspace.add_chart(chart, w37)

    column += 10

    
workspace.freeze_panes = "B2"

workbook.save(file_excel)

table = []
for cpu in cpu_list:
    cpu = '_cpu' + cpu
    cache_impact_name = 'cache_impact' + cpu
    cpi_name = 'cpi' + cpu
    for index, row in data.iterrows():
        if cache_impact_name in row:
            cache_impact = row[cache_impact_name]
            table.append({'time': index, 'cpu': cpu, 'kpi': 'cache_impact', 'value': cache_impact})
        if cpi_name in row:
            cpi = row[cpi_name]
            table.append({'time': index, 'cpu': cpu, 'kpi': 'cpi', 'value': cpi})
table = pd.DataFrame(table)
print(table)

from plotly.offline import iplot
from plotly.subplots import make_subplots
import plotly.express as px
if 'time' in table.columns:
    fig = px.line(table, x='time', y = 'value', color = 'cpu', facet_row = 'kpi')

    file_image = filename.replace('.head','.png')
    fig.write_image(file_image)

